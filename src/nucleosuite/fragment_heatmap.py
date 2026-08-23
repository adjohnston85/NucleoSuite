#!/usr/bin/env python3
"""Create clustered heatmaps from NucleoSuite fragment-length tables.

The command supports whole-genome profiles from multiple samples and multiple
BED/chromatin-state profiles from one or more samples.

Recognised long-format columns
------------------------------
Required:
    fragment_length (or frag_len/length/insert_size)
    count (or fragment_count/fragments/frequency)
Optional:
    sample
    label (or state/category/region/bed_label)

Two-column tables (fragment_length, count) and wide tables with one profile per
count column are also accepted.

Default heatmap transformation
------------------------------
1. Convert every profile to percentages across the selected fragment range.
2. For every fragment length, z-score those percentages across profiles.
3. Plot with a blue-white-orange diverging colour map centred at zero.

Outputs derived from --out-prefix:
    <prefix>_heatmap.png
    <prefix>_clustered_profiles.tsv
    <prefix>_clustered_fragment_stats.tsv
    <prefix>_normalised_matrix.tsv
    <prefix>_heatmap_plot_metadata.tsv
    <prefix>_heatmap_linkage.tsv
Optional with --write-detail-tables:
    <prefix>.xlsx
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Mapping, Optional, Sequence, Set, Tuple

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap, ListedColormap, Normalize, TwoSlopeNorm
from matplotlib.patches import Patch
from nucleosuite.progress import ProgressReporter


DEFAULT_LOW_COLOUR = "#2166AC"
DEFAULT_MID_COLOUR = "#FFFFFF"
DEFAULT_HIGH_COLOUR = "#F58518"

DEFAULT_CATEGORY_COLOURS = OrderedDict(
    [
        ("Non-cancer", "#4C78A8"),
        ("Cancer", "#F58518"),
        ("Healthy", "#54A24B"),
    ]
)

SAMPLE_COLUMNS = ("sample", "sample_name")
LABEL_COLUMNS = ("label", "state", "category", "region", "region_label", "bed_label")
LENGTH_COLUMNS = (
    "fragment_length", "fragment_len", "frag_length", "frag_len",
    "length", "insert_size", "insert_length",
)
COUNT_COLUMNS = ("count", "fragments", "fragment_count", "n", "frequency")


@dataclass
class Profile:
    name: str
    counts: Dict[int, float]
    source_file: str
    sample: str = ""
    label: str = ""


@dataclass
class MetadataRecord:
    category: str = ""


def natural_key(value: str) -> List[object]:
    return [int(x) if x.isdigit() else x.lower() for x in re.split(r"(\d+)", value)]


def normalise_heading(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def find_column(fieldnames: Sequence[str], candidates: Sequence[str]) -> Optional[str]:
    lookup = {normalise_heading(field): field for field in fieldnames}
    for candidate in candidates:
        if candidate in lookup:
            return lookup[candidate]
    return None


def sanitise_sheet_name(name: str) -> str:
    clean = re.sub(r'[:\\/?*\[\]]', "_", name).strip() or "Profile"
    return clean[:31]


def sniff_delimiter(path: Path) -> Optional[str]:
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "\t" in line:
                return "\t"
            if "," in line:
                return ","
            return None
    return None


def read_lines(path: Path) -> List[str]:
    lines: List[str] = []
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            line = line.strip()
            if line and not line.startswith("#"):
                lines.append(line)
    return lines


def split_line(line: str, delimiter: Optional[str]) -> List[str]:
    return line.split() if delimiter is None else next(csv.reader([line], delimiter=delimiter))


def is_number(value: str) -> bool:
    try:
        return np.isfinite(float(value))
    except ValueError:
        return False


def parse_length(value: str, path: Path, line_number: int) -> int:
    try:
        number = float(value)
    except ValueError as exc:
        raise SystemExit(f"Invalid fragment length in {path} line {line_number}: {value!r}") from exc
    if not np.isfinite(number) or not number.is_integer():
        raise SystemExit(f"Fragment length must be an integer in {path} line {line_number}: {value!r}")
    return int(number)


def parse_count(value: str, path: Path, line_number: int) -> float:
    try:
        number = float(value)
    except ValueError as exc:
        raise SystemExit(f"Invalid count in {path} line {line_number}: {value!r}") from exc
    if not np.isfinite(number) or number < 0:
        raise SystemExit(f"Count must be finite and non-negative in {path} line {line_number}: {value!r}")
    return number


def profile_name(sample: str, label: str, fallback: str, separator: str) -> str:
    sample = sample.strip()
    label = label.strip()
    if sample and label:
        return f"{sample}{separator}{label}"
    return sample or label or fallback


def parse_long_table(
    path: Path,
    lines: Sequence[str],
    delimiter: Optional[str],
    separator: str,
    forced_name: Optional[str],
) -> Optional[List[Profile]]:
    header = split_line(lines[0], delimiter)
    length_col = find_column(header, LENGTH_COLUMNS)
    count_col = find_column(header, COUNT_COLUMNS)
    if length_col is None or count_col is None:
        return None

    sample_col = find_column(header, SAMPLE_COLUMNS)
    label_col = find_column(header, LABEL_COLUMNS)
    grouped = defaultdict(lambda: defaultdict(float))

    reader = csv.DictReader(lines, delimiter=delimiter or "\t")
    for line_number, row in enumerate(reader, start=2):
        raw_length = (row.get(length_col) or "").strip()
        raw_count = (row.get(count_col) or "").strip()
        if not raw_length or not raw_count:
            continue

        sample = (row.get(sample_col) or "").strip() if sample_col else ""
        label = (row.get(label_col) or "").strip() if label_col else ""
        name = forced_name if forced_name and not sample and not label else profile_name(
            sample, label, forced_name or path.stem, separator
        )
        length = parse_length(raw_length, path, line_number)
        count = parse_count(raw_count, path, line_number)
        grouped[(name, sample, label)][length] += count

    return [
        Profile(name, dict(counts), str(path), sample, label)
        for (name, sample, label), counts in grouped.items()
    ]


def parse_simple_or_wide_table(
    path: Path,
    lines: Sequence[str],
    delimiter: Optional[str],
    forced_name: Optional[str],
) -> List[Profile]:
    first = split_line(lines[0], delimiter)
    header_present = not (len(first) >= 2 and is_number(first[0]) and is_number(first[1]))

    if header_present:
        header = first
        data_lines = lines[1:]
        start_line = 2
    else:
        header = ["fragment_length", forced_name or path.stem]
        data_lines = lines
        start_line = 1

    if len(header) < 2:
        raise SystemExit(f"Expected at least two columns in {path}")

    length_heading = find_column(header, LENGTH_COLUMNS)
    length_index = header.index(length_heading) if length_heading else 0
    value_indices = [i for i in range(len(header)) if i != length_index]

    names: Dict[int, str] = {}
    counts = {}
    for index in value_indices:
        if forced_name and len(value_indices) == 1:
            name = forced_name
        elif header_present:
            name = header[index].strip() or f"{path.stem}_{index + 1}"
        else:
            name = path.stem
        names[index] = name
        counts[index] = defaultdict(float)

    for line_number, line in enumerate(data_lines, start=start_line):
        fields = split_line(line, delimiter)
        if len(fields) <= length_index:
            continue
        length = parse_length(fields[length_index], path, line_number)
        for index in value_indices:
            if index >= len(fields) or not fields[index].strip():
                continue
            counts[index][length] += parse_count(fields[index], path, line_number)

    return [Profile(names[i], dict(counts[i]), str(path)) for i in value_indices]


def read_profiles(path: Path, separator: str, forced_name: Optional[str]) -> List[Profile]:
    if not path.is_file():
        raise SystemExit(f"Input file not found: {path}")
    lines = read_lines(path)
    if not lines:
        raise SystemExit(f"Input file is empty: {path}")
    delimiter = sniff_delimiter(path)
    profiles = parse_long_table(path, lines, delimiter, separator, forced_name)
    if profiles is None:
        profiles = parse_simple_or_wide_table(path, lines, delimiter, forced_name)
    if not profiles:
        raise SystemExit(f"No profiles found in {path}")
    return profiles


def parse_inputs(values: Sequence[str]) -> List[Tuple[Optional[str], Path]]:
    result = []
    for value in values:
        if "=" in value:
            name, filename = value.split("=", 1)
            if name.strip() and Path(filename).exists():
                result.append((name.strip(), Path(filename)))
                continue
        result.append((None, Path(value)))
    return result


def make_unique(profiles: List[Profile], policy: str) -> List[Profile]:
    seen: Dict[str, int] = {}
    for profile in profiles:
        seen[profile.name] = seen.get(profile.name, 0) + 1
        if seen[profile.name] > 1:
            if policy == "error":
                raise SystemExit(
                    f"Duplicate profile name {profile.name!r}. Use NAME=FILE or --duplicate-policy suffix."
                )
            profile.name = f"{profile.name}_{seen[profile.name]}"
    return profiles


def dense_matrix(profiles: Sequence[Profile], min_frag: int, max_frag: int) -> Tuple[np.ndarray, np.ndarray]:
    lengths = np.arange(min_frag, max_frag + 1, dtype=int)
    matrix = np.zeros((len(profiles), len(lengths)), dtype=float)
    for row, profile in enumerate(profiles):
        for length, count in profile.counts.items():
            if min_frag <= length <= max_frag:
                matrix[row, length - min_frag] += count
    return matrix, lengths


def row_percent(matrix: np.ndarray) -> np.ndarray:
    totals = matrix.sum(axis=1, keepdims=True)
    return np.divide(matrix, totals, out=np.zeros_like(matrix, dtype=float), where=totals > 0) * 100.0


def column_percent(matrix: np.ndarray) -> np.ndarray:
    totals = matrix.sum(axis=0, keepdims=True)
    return np.divide(matrix, totals, out=np.zeros_like(matrix, dtype=float), where=totals > 0) * 100.0


def zscore_columns(matrix: np.ndarray) -> np.ndarray:
    means = matrix.mean(axis=0, keepdims=True)
    stds = matrix.std(axis=0, ddof=0, keepdims=True)
    return np.divide(matrix - means, stds, out=np.zeros_like(matrix), where=stds > 0)


def minmax(matrix: np.ndarray, axis: int) -> np.ndarray:
    minima = matrix.min(axis=axis, keepdims=True)
    ranges = matrix.max(axis=axis, keepdims=True) - minima
    return np.divide(matrix - minima, ranges, out=np.zeros_like(matrix), where=ranges > 0)


def normalise(matrix: np.ndarray, method: str) -> Tuple[np.ndarray, str, Optional[float]]:
    clean = np.asarray(matrix, dtype=float).copy()
    clean[~np.isfinite(clean)] = 0.0
    clean = np.clip(clean, 0.0, None)

    if method == "fragment-zscore":
        return zscore_columns(row_percent(clean)), "Fragment-length z-score across profiles", 0.0
    if method == "profile-percent":
        return row_percent(clean), "Fragments within profile (%)", None
    if method == "fragment-percent":
        return column_percent(clean), "Profiles within fragment length (%)", None
    if method == "profile-minmax":
        return minmax(clean, axis=1), "Profile-scaled value (0-1)", None
    if method == "fragment-minmax":
        return minmax(clean, axis=0), "Fragment-length-scaled value (0-1)", None
    return clean, "Fragment count", None


def downsample(matrix: np.ndarray, target: Optional[str], seed: int) -> Tuple[np.ndarray, Optional[int]]:
    if target is None:
        return matrix.copy(), None

    integer_matrix = np.rint(np.clip(matrix, 0, None)).astype(np.int64)
    totals = integer_matrix.sum(axis=1)
    nonzero = totals[totals > 0]
    if nonzero.size == 0:
        raise SystemExit("All profiles are empty within the selected range.")

    if target.lower() in {"min", "minimum", "fewest"}:
        target_n = int(nonzero.min())
    else:
        try:
            target_n = int(target)
        except ValueError as exc:
            raise SystemExit("--downsample-to must be a positive integer or 'min'") from exc
        if target_n <= 0:
            raise SystemExit("--downsample-to must be greater than zero")

    rng = np.random.default_rng(seed)
    result = integer_matrix.copy()
    for row, total in enumerate(totals):
        total = int(total)
        if total > target_n:
            result[row] = rng.multinomial(target_n, result[row] / float(total))
    return result.astype(float), target_n


def cluster(matrix: np.ndarray, method: str, metric: str) -> Tuple[np.ndarray, Optional[np.ndarray]]:
    if matrix.shape[0] < 2:
        return np.arange(matrix.shape[0]), None
    from scipy.cluster.hierarchy import dendrogram, linkage
    from scipy.spatial.distance import pdist
    distances = pdist(matrix, metric=metric)
    if not np.all(np.isfinite(distances)):
        raise ValueError("The selected metric produced non-finite distances.")
    linkage_matrix = linkage(distances, method=method)
    order = np.asarray(dendrogram(linkage_matrix, no_plot=True)["leaves"], dtype=int)
    return order, linkage_matrix


def read_metadata(path: Optional[Path], profile_column: str, category_column: str) -> Dict[str, MetadataRecord]:
    if path is None:
        return {}
    if not path.is_file():
        raise SystemExit(f"Metadata file not found: {path}")

    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        if not reader.fieldnames:
            raise SystemExit(f"Metadata file has no header: {path}")
        lookup = {normalise_heading(field): field for field in reader.fieldnames}
        profile_key = lookup.get(normalise_heading(profile_column))
        category_key = lookup.get(normalise_heading(category_column))
        if profile_key is None:
            raise SystemExit(f"Metadata profile column {profile_column!r} not found")
        records = {}
        for row in reader:
            name = (row.get(profile_key) or "").strip()
            if name:
                category = (row.get(category_key) or "").strip() if category_key else ""
                records[name] = MetadataRecord(category)
        return records


def parse_category_colours(values: Sequence[str]) -> OrderedDict[str, str]:
    colours = OrderedDict(DEFAULT_CATEGORY_COLOURS)
    for value in values:
        if "=" not in value:
            raise SystemExit("--category-colour must use CATEGORY=COLOUR")
        category, colour = value.split("=", 1)
        if not category.strip() or not colour.strip():
            raise SystemExit("--category-colour must use CATEGORY=COLOUR")
        colours[category.strip()] = colour.strip()
    return colours


def observed_category_colours(categories: Sequence[str], requested: Mapping[str, str]) -> OrderedDict[str, str]:
    lower = {name.lower(): colour for name, colour in requested.items()}
    fallbacks = ["#72B7B2", "#B279A2", "#FF9DA6", "#9D755D", "#BAB0AC", "#E45756"]
    result: OrderedDict[str, str] = OrderedDict()
    fallback_index = 0
    for category in categories:
        if not category or category in result:
            continue
        colour = lower.get(category.lower())
        if colour is None:
            colour = fallbacks[fallback_index % len(fallbacks)]
            fallback_index += 1
        result[category] = colour
    return result


def colour_limits(
    matrix: np.ndarray,
    centre: Optional[float],
    percentile: float,
    explicit_min: Optional[float],
    explicit_max: Optional[float],
) -> Tuple[float, float]:
    finite = matrix[np.isfinite(matrix)]
    if finite.size == 0:
        return -1.0, 1.0
    vmin = float(explicit_min) if explicit_min is not None else float(np.percentile(finite, 100 - percentile))
    vmax = float(explicit_max) if explicit_max is not None else float(np.percentile(finite, percentile))
    if centre is not None:
        extent = max(abs(vmin - centre), abs(vmax - centre), 1e-12)
        return centre - extent, centre + extent
    if vmax <= vmin:
        pad = abs(vmax) * 0.05 or 1.0
        return vmin - pad, vmax + pad
    return vmin, vmax


def make_cmap(name: Optional[str], low: str, middle: str, high: str):
    if name:
        try:
            return plt.get_cmap(name)
        except ValueError as exc:
            raise SystemExit(f"Unknown Matplotlib colour map: {name}") from exc
    return LinearSegmentedColormap.from_list("blue_white_orange", [low, middle, high], N=256)


def save_heatmap(
    matrix: np.ndarray,
    profiles: Sequence[Profile],
    lengths: np.ndarray,
    order: np.ndarray,
    linkage_matrix: Optional[np.ndarray],
    metadata: Mapping[str, MetadataRecord],
    category_colours: Mapping[str, str],
    output: Path,
    colourbar_label: str,
    args: argparse.Namespace,
    centre: Optional[float],
) -> None:
    clustered = matrix[order]
    names = [profiles[i].name for i in order]
    categories = [metadata.get(name, MetadataRecord()).category for name in names]
    have_categories = any(categories)
    category_map = observed_category_colours(categories, category_colours)

    height = max(4.5, min(0.24 * len(names) + 2.5, 24.0))
    width = max(14.0, min(0.045 * len(lengths) + 12.0, 40.0))
    figure = plt.figure(figsize=(width, height))

    ratios = []
    if linkage_matrix is not None:
        ratios.append(2.2)
    if have_categories:
        ratios.append(0.18)
    ratios.extend([args.label_gutter, 10.0, 0.38])
    grid = figure.add_gridspec(1, len(ratios), width_ratios=ratios, wspace=0.08)
    column = 0

    if linkage_matrix is not None:
        from scipy.cluster.hierarchy import dendrogram
        axis = figure.add_subplot(grid[0, column])
        column += 1
        dendrogram(
            linkage_matrix,
            orientation="left",
            no_labels=True,
            color_threshold=0,
            above_threshold_color=args.dendrogram_colour,
            link_color_func=lambda _: args.dendrogram_colour,
            ax=axis,
        )
        axis.invert_yaxis()
        axis.axis("off")

    if have_categories:
        axis = figure.add_subplot(grid[0, column])
        column += 1
        keys = list(category_map)
        codes = {name: i for i, name in enumerate(keys)}
        values = np.asarray([codes.get(category, 0) for category in categories])
        axis.imshow(
            values.reshape(-1, 1), aspect="auto", interpolation="nearest", origin="upper",
            cmap=ListedColormap([category_map[key] for key in keys]),
            vmin=-0.5, vmax=max(len(keys) - 0.5, 0.5),
        )
        axis.set_xticks([])
        axis.set_yticks([])
        axis.set_title("Category", fontsize=9, pad=6)

    gutter = figure.add_subplot(grid[0, column])
    column += 1
    gutter.axis("off")
    heatmap_axis = figure.add_subplot(grid[0, column])
    column += 1
    colourbar_axis = figure.add_subplot(grid[0, column])

    vmin, vmax = colour_limits(clustered, centre, args.colour_percentile, args.colour_min, args.colour_max)
    norm = TwoSlopeNorm(vmin=vmin, vcenter=centre, vmax=vmax) if centre is not None and vmin < centre < vmax else Normalize(vmin=vmin, vmax=vmax)
    image = heatmap_axis.imshow(
        clustered, aspect="auto", interpolation="nearest", origin="upper",
        cmap=make_cmap(args.heatmap_cmap, args.low_colour, args.mid_colour, args.high_colour),
        norm=norm,
    )

    heatmap_axis.set_xlabel("Fragment length (bp)")
    heatmap_axis.grid(False)
    step = 1 if len(names) <= args.max_yticks else int(np.ceil(len(names) / args.max_yticks))
    y_positions = np.arange(0, len(names), step)
    heatmap_axis.set_yticks(y_positions)
    heatmap_axis.set_yticklabels([names[i] for i in y_positions], fontsize=8)
    heatmap_axis.tick_params(axis="y", length=0, pad=6)

    first_tick = int(math.ceil(lengths[0] / 50.0) * 50)
    tick_values = np.arange(first_tick, lengths[-1] + 1, 50, dtype=int)
    heatmap_axis.set_xticks(tick_values - lengths[0])
    heatmap_axis.set_xticklabels([str(x) for x in tick_values], rotation=45, ha="right", fontsize=8)
    if args.title:
        heatmap_axis.set_title(args.title, pad=18)

    colourbar = figure.colorbar(image, cax=colourbar_axis)
    colourbar.set_label(colourbar_label)
    colourbar.ax.tick_params(labelsize=8)

    if have_categories and category_map:
        handles = [Patch(facecolor=colour, edgecolor="none", label=name) for name, colour in category_map.items()]
        heatmap_axis.legend(
            handles=handles, loc="lower left", bbox_to_anchor=(0.0, 1.005),
            ncol=min(4, len(handles)), frameon=False, fontsize=9,
        )

    figure.text(0.01, 0.5, "Profile", rotation=90, va="center", fontsize=10)
    plt.subplots_adjust(left=0.06, right=0.95, top=0.91, bottom=0.12, wspace=0.08)
    from nucleosuite.plotting import save_figure
    save_figure(figure, output, default_dpi=args.dpi, bbox_inches="tight")
    plt.close(figure)


def weighted_median_mode(counts: np.ndarray, lengths: np.ndarray) -> Tuple[Optional[int], Optional[int]]:
    total = counts.sum()
    if total <= 0:
        return None, None
    median_index = int(np.searchsorted(np.cumsum(counts), total * 0.5, side="left"))
    mode_index = int(np.argmax(counts))
    return int(lengths[median_index]), int(lengths[mode_index])


def write_profile_order(path: Path, profiles: Sequence[Profile], order: np.ndarray, metadata: Mapping[str, MetadataRecord]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow(["cluster_order", "original_index", "profile", "sample", "label", "category", "source_file"])
        for position, index in enumerate(order, start=1):
            p = profiles[index]
            writer.writerow([position, int(index), p.name, p.sample, p.label, metadata.get(p.name, MetadataRecord()).category, p.source_file])


def write_plot_metadata(
    path: Path,
    *,
    args: argparse.Namespace,
    colourbar_label: str,
    centre: Optional[float],
    category_colours: Mapping[str, str],
    clustered: bool,
) -> None:
    """Write the presentation settings needed for an exact ``plot`` replot."""
    rows = [
        ("normalisation", args.normalisation),
        ("colourbar_label", colourbar_label),
        ("heatmap_cmap", args.heatmap_cmap or ""),
        ("low_colour", args.low_colour),
        ("mid_colour", args.mid_colour),
        ("high_colour", args.high_colour),
        ("heatmap_centre", "" if centre is None else centre),
        ("colour_percentile", args.colour_percentile),
        ("colour_min", "" if args.colour_min is None else args.colour_min),
        ("colour_max", "" if args.colour_max is None else args.colour_max),
        ("dendrogram_colour", args.dendrogram_colour),
        ("label_gutter", args.label_gutter),
        ("max_yticks", args.max_yticks),
        ("title", args.title or ""),
        ("clustered", "true" if clustered else "false"),
    ]
    rows.extend((f"category_colour:{name}", colour) for name, colour in category_colours.items())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow(["setting", "value"])
        writer.writerows(rows)


def write_linkage(path: Path, linkage_matrix: Optional[np.ndarray]) -> None:
    """Write SciPy linkage rows so a replot can retain the original dendrogram."""
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow(["left_child", "right_child", "distance", "member_count"])
        if linkage_matrix is not None:
            for left, right, distance, count in linkage_matrix:
                writer.writerow([int(left), int(right), f"{distance:.17g}", int(count)])


def write_stats(
    path: Path,
    profiles: Sequence[Profile],
    original: np.ndarray,
    analysed: np.ndarray,
    lengths: np.ndarray,
    order: np.ndarray,
    metadata: Mapping[str, MetadataRecord],
) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow([
            "profile", "sample", "label", "category", "original_total_in_range",
            "analysed_total_in_range", "median_frag_bp", "mode_frag_bp", "source_file",
        ])
        for index in order:
            p = profiles[index]
            median, mode = weighted_median_mode(analysed[index], lengths)
            writer.writerow([
                p.name, p.sample, p.label, metadata.get(p.name, MetadataRecord()).category,
                int(round(original[index].sum())), int(round(analysed[index].sum())),
                "" if median is None else median, "" if mode is None else mode, p.source_file,
            ])


def write_matrix(path: Path, matrix: np.ndarray, profiles: Sequence[Profile], lengths: np.ndarray, order: np.ndarray) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow(["profile", *lengths.tolist()])
        for index in order:
            writer.writerow([profiles[index].name, *[f"{x:.10g}" for x in matrix[index]]])


def add_excel_chart(worksheet, title: str, min_frag: int, max_frag: int) -> None:
    from openpyxl.chart import Reference, ScatterChart, Series
    from openpyxl.chart.marker import Marker
    x_values = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
    y_values = Reference(worksheet, min_col=3, min_row=2, max_row=worksheet.max_row)
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = "Fragment length (bp)"
    chart.y_axis.title = "Profile percent"
    chart.x_axis.scaling.min = min_frag
    chart.x_axis.scaling.max = max_frag
    chart.x_axis.majorUnit = 50
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = None
    chart.legend = None
    series = Series(y_values, x_values, title=None)
    series.smooth = True
    series.marker = Marker(symbol="none")
    chart.series.append(series)
    worksheet.add_chart(chart, "G2")


def write_excel(
    path: Path,
    profiles: Sequence[Profile],
    original: np.ndarray,
    analysed: np.ndarray,
    display: np.ndarray,
    lengths: np.ndarray,
    order: np.ndarray,
    metadata: Mapping[str, MetadataRecord],
    args: argparse.Namespace,
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise SystemExit("Excel output requires openpyxl; install it or use --no-excel") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    used: Set[str] = set()
    percentages = row_percent(analysed)

    for index in order:
        p = profiles[index]
        base = sanitise_sheet_name(p.name)
        sheet = base
        counter = 2
        while sheet in used:
            suffix = f"_{counter}"
            sheet = sanitise_sheet_name(base[:31 - len(suffix)] + suffix)
            counter += 1
        used.add(sheet)
        worksheet = workbook.create_sheet(sheet)
        worksheet.append(["fragment_length", "original_count", "profile_percent", "analysed_count", "heatmap_value"])
        for j, length in enumerate(lengths):
            worksheet.append([int(length), float(original[index, j]), float(percentages[index, j]), float(analysed[index, j]), float(display[index, j])])
        worksheet["I2"] = "profile"
        worksheet["J2"] = p.name
        worksheet["I3"] = "sample"
        worksheet["J3"] = p.sample
        worksheet["I4"] = "label"
        worksheet["J4"] = p.label
        worksheet["I5"] = "category"
        worksheet["J5"] = metadata.get(p.name, MetadataRecord()).category
        worksheet["I6"] = "source_file"
        worksheet["J6"] = p.source_file
        worksheet["I7"] = "heatmap_normalisation"
        worksheet["J7"] = args.normalisation
        add_excel_chart(worksheet, f"{p.name}: fragment-length profile", args.min_frag, args.max_frag)
    workbook.save(path)


def parse_exclusions(direct: Sequence[str], path: Optional[Path]) -> Set[str]:
    result = {x.strip() for x in direct if x.strip()}
    if path:
        if not path.is_file():
            raise SystemExit(f"Exclusion file not found: {path}")
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for line in handle:
                line = line.strip()
                if line and not line.startswith("#"):
                    result.add(line)
    return result


def add_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("-i", "--input", action="append", required=True, metavar="[NAME=]FILE", help="Input table; repeat for multiple files.")
    parser.add_argument("-o", "--out-prefix", help="Output path prefix. Default: first input-table basename plus _fragment_heatmap.")
    parser.add_argument("--min-frag", type=int, default=1, help="Minimum fragment length. Default: 1")
    parser.add_argument("--max-frag", type=int, default=500, help="Maximum fragment length. Default: 500")
    parser.add_argument(
        "--normalization", "--normalisation", dest="normalisation",
        choices=["fragment-zscore", "profile-percent", "fragment-percent", "profile-minmax", "fragment-minmax", "none"],
        default="fragment-zscore",
        help="Heatmap normalisation. Default: fragment-zscore",
    )
    parser.add_argument("--downsample-to", default=None, metavar="N|min", help="Optional multinomial downsampling target.")
    parser.add_argument("--seed", type=int, default=1, help="Downsampling seed. Default: 1")
    parser.add_argument("--cluster-method", choices=["average", "complete", "single", "weighted", "centroid", "median", "ward"], default="average", help="Hierarchical-linkage method used to order profiles (default: average).")
    parser.add_argument("--cluster-metric", default="euclidean", help="SciPy distance metric used for profile clustering (default: euclidean).")
    parser.add_argument("--no-cluster", action="store_true", help="Keep input profile order and omit the dendrogram.")
    parser.add_argument("--metadata", type=Path, default=None, help="Optional tab-delimited profile metadata.")
    parser.add_argument("--metadata-profile-column", default="profile", help="Metadata column matched to profile names (default: profile).")
    parser.add_argument("--metadata-category-column", default="category", help="Metadata column used for category annotations (default: category).")
    parser.add_argument("--require-metadata", action="store_true", help="Stop if any retained profile lacks a metadata category.")
    parser.add_argument("--category-colour", action="append", default=[], metavar="CATEGORY=COLOUR", help="Assign a Matplotlib colour to one metadata category; repeat as needed.")
    parser.add_argument("--heatmap-cmap", default=None, help="Named Matplotlib cmap; gray_r restores grayscale.")
    parser.add_argument("--low-colour", default=DEFAULT_LOW_COLOUR, help=f"Low-value colour for the custom heatmap palette (default: {DEFAULT_LOW_COLOUR}).")
    parser.add_argument("--mid-colour", default=DEFAULT_MID_COLOUR, help=f"Midpoint colour for the custom heatmap palette (default: {DEFAULT_MID_COLOUR}).")
    parser.add_argument("--high-colour", default=DEFAULT_HIGH_COLOUR, help=f"High-value colour for the custom heatmap palette (default: {DEFAULT_HIGH_COLOUR}).")
    parser.add_argument("--heatmap-centre", type=float, default=None, help="Explicit colour-map centre; default: 0 for z-scores and automatic otherwise.")
    parser.add_argument("--colour-percentile", type=float, default=99.0, help="Symmetric colour-limit percentile when explicit limits are absent (default: 99).")
    parser.add_argument("--colour-min", type=float, default=None, help="Explicit heatmap colour minimum.")
    parser.add_argument("--colour-max", type=float, default=None, help="Explicit heatmap colour maximum.")
    parser.add_argument("--dendrogram-colour", default="black", help="Colour used for every dendrogram branch (default: black).")
    parser.add_argument("--profile-separator", default="::", help="Separator joining source and subgroup names from long-format tables (default: ::).")
    parser.add_argument("--duplicate-policy", choices=["error", "suffix"], default="error", help="Handle duplicate profile names by stopping or adding numeric suffixes (default: error).")
    parser.add_argument("--exclude-profile", action="append", default=[], help="Profile name to exclude; repeat for multiple profiles.")
    parser.add_argument("--exclude-profiles-file", type=Path, default=None, help="Text file containing one profile name to exclude per line.")
    parser.add_argument("--label-gutter", type=float, default=1.2, help="Relative width reserved for row labels (default: 1.2).")
    parser.add_argument("--max-yticks", type=int, default=80, help="Maximum number of displayed profile labels before tick thinning (default: 80).")
    parser.add_argument("--dpi", type=int, default=220, help="Heatmap resolution setting (default: 220); --plot-dpi takes precedence when supplied.")
    parser.add_argument("--title", default=None, help="Optional heatmap title.")
    parser.add_argument(
        "--write-detail-tables", action="store_true",
        help=(
            "Write the full Excel workbook in addition to the standard plot and "
            "normalized matrix. The matrix remains a default compact plot source "
            "so the heatmap can always be recreated with `nucleosuite plot`."
        ),
    )
    parser.add_argument("--no-excel", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--no-matrix", action="store_true", help=argparse.SUPPRESS)


def run(args: argparse.Namespace) -> int:
    if not args.out_prefix:
        from nucleosuite.output_naming import automatic_prefix
        first = args.input[0].split("=", 1)[1] if "=" in args.input[0] else args.input[0]
        args.out_prefix = str(automatic_prefix(first, "fragment_heatmap"))
    if args.min_frag < 0 or args.max_frag < args.min_frag:
        raise SystemExit("Require 0 <= --min-frag <= --max-frag")
    if not 50 < args.colour_percentile <= 100:
        raise SystemExit("--colour-percentile must be >50 and <=100")
    if args.cluster_method == "ward" and args.cluster_metric != "euclidean":
        raise SystemExit("Ward clustering requires Euclidean distance")

    reporter = ProgressReporter("fragment-heatmap")
    reporter.stage("Loading fragment-length profiles")
    profiles: List[Profile] = []
    for forced_name, path in parse_inputs(args.input):
        profiles.extend(read_profiles(path, args.profile_separator, forced_name))
    profiles = make_unique(profiles, args.duplicate_policy)

    exclusions = parse_exclusions(args.exclude_profile, args.exclude_profiles_file)
    profiles = [p for p in profiles if p.name not in exclusions]
    profiles.sort(key=lambda p: natural_key(p.name))
    if not profiles:
        raise SystemExit("No profiles remain after exclusions")

    reporter.stage(
        f"Building fragment-length matrix for {len(profiles):,} profiles"
    )
    original, lengths = dense_matrix(profiles, args.min_frag, args.max_frag)
    keep = original.sum(axis=1) > 0
    if not np.all(keep):
        empty = [p.name for p, include in zip(profiles, keep) if not include]
        print("Warning: excluding empty profiles: " + ", ".join(empty), file=sys.stderr)
        profiles = [p for p, include in zip(profiles, keep) if include]
        original = original[keep]
    if not profiles:
        raise SystemExit("No profiles contain fragments within the selected range")

    analysed, target_n = downsample(original, args.downsample_to, args.seed)
    if target_n is not None:
        print(f"Downsampling target: {target_n:,} fragments per profile (seed={args.seed})")

    display, colourbar_label, default_centre = normalise(analysed, args.normalisation)
    centre = args.heatmap_centre if args.heatmap_centre is not None else default_centre
    metadata = read_metadata(args.metadata, args.metadata_profile_column, args.metadata_category_column)
    if args.require_metadata:
        missing = [p.name for p in profiles if p.name not in metadata]
        if missing:
            raise SystemExit("Profiles missing from metadata: " + ", ".join(missing))

    reporter.stage("Normalizing and ordering heatmap rows")
    if args.no_cluster:
        order, linkage_matrix = np.arange(len(profiles)), None
    else:
        try:
            order, linkage_matrix = cluster(display, args.cluster_method, args.cluster_metric)
        except Exception as exc:
            print(f"Warning: clustering failed; retaining natural order. Reason: {exc}", file=sys.stderr)
            order, linkage_matrix = np.arange(len(profiles)), None

    from nucleosuite.output_naming import parameterized_prefix

    prefix = parameterized_prefix(
        args.out_prefix,
        (
            ("fragmin", args.min_frag),
            ("fragmax", args.max_frag),
            ("norm", args.normalisation),
            ("cluster", "none" if args.no_cluster else f"{args.cluster_method}-{args.cluster_metric}"),
            ("downsample", args.downsample_to),
        ),
    )
    prefix.parent.mkdir(parents=True, exist_ok=True)
    from nucleosuite.plotting import plot_path
    heatmap_path = plot_path(Path(f"{prefix}_heatmap.png"))
    profile_path = Path(f"{prefix}_clustered_profiles.tsv")
    stats_path = Path(f"{prefix}_clustered_fragment_stats.tsv")
    matrix_path = Path(f"{prefix}_normalised_matrix.tsv")
    plot_metadata_path = Path(f"{prefix}_heatmap_plot_metadata.tsv")
    linkage_path = Path(f"{prefix}_heatmap_linkage.tsv")
    excel_path = Path(f"{prefix}.xlsx")

    reporter.stage("Writing heatmap and tables")
    requested_category_colours = parse_category_colours(args.category_colour)
    save_heatmap(
        display, profiles, lengths, order, linkage_matrix, metadata,
        requested_category_colours, heatmap_path,
        colourbar_label, args, centre,
    )
    print(f"Wrote: {heatmap_path}")
    write_profile_order(profile_path, profiles, order, metadata)
    print(f"Wrote: {profile_path}")
    write_stats(stats_path, profiles, original, analysed, lengths, order, metadata)
    print(f"Wrote: {stats_path}")
    write_plot_metadata(
        plot_metadata_path,
        args=args,
        colourbar_label=colourbar_label,
        centre=centre,
        category_colours=requested_category_colours,
        clustered=linkage_matrix is not None,
    )
    print(f"Wrote: {plot_metadata_path}")
    write_linkage(linkage_path, linkage_matrix)
    print(f"Wrote: {linkage_path}")
    if args.no_matrix:
        print(
            "WARNING: --no-matrix is deprecated and ignored; the normalized matrix "
            "is retained as the compact replot source.",
            file=sys.stderr,
        )
    write_matrix(matrix_path, display, profiles, lengths, order)
    print(f"Wrote: {matrix_path}")
    if args.write_detail_tables and not args.no_excel:
        write_excel(excel_path, profiles, original, analysed, display, lengths, order, metadata, args)
        print(f"Wrote: {excel_path}")
    return 0


def register_parser(subparsers) -> argparse.ArgumentParser:
    parser = subparsers.add_parser(
        "fragment-heatmap",
        help="Create clustered heatmaps from fragment-length profiles.",
    )
    add_arguments(parser)
    parser.set_defaults(command_function=run)
    from nucleosuite.plotting import add_plotting_arguments
    add_plotting_arguments(parser)
    return parser


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="nucleosuite fragment-heatmap", description=__doc__)
    add_arguments(parser)
    from nucleosuite.plotting import add_plotting_arguments
    add_plotting_arguments(parser)
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    return run(build_parser().parse_args(argv))


if __name__ == "__main__":
    raise SystemExit(main())
